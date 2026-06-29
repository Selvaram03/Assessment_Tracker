"""
helpers.py

Utility functions for Assessment Automation
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from copy import copy

import pandas as pd
from openpyxl.utils import get_column_letter

from config import *


# ==========================================================
# Filename Parser
# ==========================================================

def parse_filename(filename: str):
    """
    Example:

    Daily_Score_Report_Java_A1_24_06_2026.xlsx
    Daily_Score_Report_All_Colleges_Assessment_117.xlsx
    Daily_Attendance_Report_All_Colleges_Assessment_117.xlsx

    Returns

    course = Java
    batch = A1
    date = 24-06-2026
    """

    name = filename.replace(".xlsx", "").replace(".xls", "")

    # Preferred naming format:
    # Daily_Score_Report_<course>_<batch or batch-group>_<dd>_<mm>_<yyyy>_<id>
    pattern = (
        r"^Daily_Score_Report_"
        r"(?P<course>.+?)_"
        r"(?P<batch>(?:[AB]\d+(?:_[AB]\d+)*))_"
        r"(?P<day>\d{2})_"
        r"(?P<month>\d{2})_"
        r"(?P<year>\d{4})_"
        r"(?P<report_id>\d+)$"
    )

    match = re.search(pattern, name)

    if match:
        course = match.group("course").replace("_", " ").strip()
        batch = match.group("batch").strip()
        assessment_date = (
            f"{match.group('day')}-"
            f"{match.group('month')}-"
            f"{match.group('year')}"
        )

        return course, batch, assessment_date

    generic_assessment_pattern = (
        r"Daily_(?:Score|Attendance)_Report_(.*?)_Assessment_(\d+)"
    )

    match = re.search(generic_assessment_pattern, name)

    if match:
        course = match.group(1).replace("_", " ").strip()
        report_id = match.group(2)

        # Keep the file in the existing flow by assigning a valid batch prefix.
        batch = f"A{report_id}"
        assessment_date = report_id

        return course, batch, assessment_date

    raise ValueError(f"Invalid filename : {filename}")


# ==========================================================
# Batch Type
# ==========================================================

def get_batch_type(batch: str):

    batch = batch.upper().strip()

    tokens = [
        token
        for token in batch.split("_")
        if token
    ]

    if not tokens:
        raise ValueError("Unknown Batch : empty")

    token_types = {
        token[0]
        for token in tokens
        if token[0] in {"A", "B"}
    }

    if token_types == {"A"}:
        return "Advanced"

    if token_types == {"B"}:
        return "Intermediate"

    if token_types == {"A", "B"}:
        return "Mixed"

    raise ValueError(f"Unknown Batch : {batch}")


# ==========================================================
# Column Letter
# ==========================================================

def column_name(column_number):

    return get_column_letter(column_number)


# ==========================================================
# Copy Cell
# ==========================================================

def copy_cell(source, target):

    target.value = source.value

    if source.has_style:

        target.font = copy(source.font)

        target.border = copy(source.border)

        target.fill = copy(source.fill)

        target.number_format = copy(source.number_format)

        target.protection = copy(source.protection)

        target.alignment = copy(source.alignment)


# ==========================================================
# Register Dictionary
# ==========================================================

def build_register_map(sheet):

    """
    Returns

    {

        Register Number : Excel Row

    }

    """

    register_map = {}

    row = DATA_START_ROW

    while True:

        value = sheet.cell(row=row, column=1).value

        if value is None:

            break

        register_map[str(value).strip()] = row

        row += 1

    return register_map


# ==========================================================
# Last Date Block
# ==========================================================

def get_last_date_column(sheet):

    """
    Detect last assessment block.

    Returns starting column
    """

    last = 2

    for merged in sheet.merged_cells.ranges:

        if merged.min_row != DATE_ROW:
            continue

        if merged.min_col > last:
            last = merged.min_col

    return last


# ==========================================================
# Competition Ranking
# ==========================================================

def apply_competition_rank(df):

    """
    Competition Ranking

    Score Desc

    Time Asc

    """

    df = df.sort_values(

        by=SORT_COLUMNS,

        ascending=SORT_ASCENDING

    ).reset_index(drop=True)

    ranks = []

    current_rank = 1

    previous_score = None

    previous_time = None

    previous_position = 1

    for position, row in enumerate(df.itertuples(index=False), start=1):

        score = getattr(row, SCORE_COLUMN.replace(" ", "_"))

        time = getattr(row, TIME_COLUMN.replace(" ", "_").replace("(", "").replace(")", ""))

        if previous_score is None:

            current_rank = 1

        elif score == previous_score and time == previous_time:

            current_rank = previous_position

        else:

            current_rank = position

        ranks.append(current_rank)

        previous_score = score

        previous_time = time

        previous_position = current_rank

    df[RANK_COLUMN] = ranks

    return df


# ==========================================================
# Percentage
# ==========================================================

def calculate_percentage(score, total):

    try:

        score = float(score)

        total = float(total)

    except (TypeError, ValueError):

        return 0

    if total == 0:
        return 0

    return round((score / total) * 100, 2)


# ==========================================================
# Status
# ==========================================================

def calculate_status(total):

    try:

        total = float(total)

    except (TypeError, ValueError):

        total = 0

    if total > 0:
        return ATTENDED

    return ABSENT


# ==========================================================
# Safe Float
# ==========================================================

def safe_float(value):

    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


# ==========================================================
# Safe Int
# ==========================================================

def safe_int(value):

    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


# ==========================================================
# Clean Register Number
# ==========================================================

def clean_register_number(value):
    """
    Handles:
    737621205001
    737621205001.0
    7.37621205001E+11
    """

    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    text = str(value).strip()

    if text == "":
        return ""

    if "e" in text.lower():
        try:
            return str(int(Decimal(text)))
        except (InvalidOperation, ValueError):
            return text

    if text.endswith(".0"):
        text = text[:-2]

    try:
        numeric = Decimal(text)
    except InvalidOperation:
        return text

    if numeric == numeric.to_integral():
        return str(int(numeric))

    return text
