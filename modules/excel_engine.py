"""
modules/excel_engine.py

Excel Processing Engine

Author : Karthi Dass
"""

from __future__ import annotations

import re
import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

from config import *
from modules.helpers import clean_register_number


class ExcelEngine:

    def __init__(self, workbook_path, master_type="IT"):

        self.workbook_path = Path(workbook_path)
        self.master_type = str(master_type).upper()
        self.is_ece = self.master_type == "ECE"

        self.workbook = None

        self.advanced_sheet = None
        self.intermediate_sheet = None
        self.main_sheet = None
        self.top_sheet = None
        self.last_sheet = None

        # Register Number -> Excel Row
        self.advanced_map = {}
        self.intermediate_map = {}
        self.main_map = {}

    # ==========================================================
    # Load Workbook
    # ==========================================================

    def load_workbook(self):

        self.workbook = load_workbook(
            self.workbook_path
        )

        self.validate_workbook_template()

        if self.is_ece:

            self.main_sheet = self.workbook.worksheets[0]

        else:

            self.advanced_sheet = self.workbook[
                ADVANCED_SHEET
            ]

            self.intermediate_sheet = self.workbook[
                INTERMEDIATE_SHEET
            ]

        if TOP10_SHEET in self.workbook.sheetnames:
            self.top_sheet = self.workbook[
                TOP10_SHEET
            ]

        if LAST10_SHEET in self.workbook.sheetnames:
            self.last_sheet = self.workbook[
                LAST10_SHEET
            ]

    # ==========================================================
    # Template Validation
    # ==========================================================

    def validate_workbook_template(self):

        if self.is_ece:
            self.validate_ece_workbook_template()
            return

        required_sheets = [
            ADVANCED_SHEET,
            INTERMEDIATE_SHEET
        ]

        missing = [
            sheet_name
            for sheet_name in required_sheets
            if sheet_name not in self.workbook.sheetnames
        ]

        if missing:
            raise ValueError(
                "Master workbook is missing required sheets: "
                + ", ".join(missing)
            )

        for sheet_name in required_sheets:
            sheet = self.workbook[sheet_name]

            has_seed_block = any(
                merged.min_row == DATE_ROW
                and merged.min_col == FIRST_ASSESSMENT_COLUMN
                and merged.max_col == FIRST_ASSESSMENT_COLUMN + DATE_BLOCK_SIZE - 1
                for merged in sheet.merged_cells.ranges
            )

            if not has_seed_block:
                raise ValueError(
                    f"{sheet_name} must contain a seed assessment block "
                    f"merged across columns G:K in row {DATE_ROW}."
                )

    def normalize_header(self, value):
        if value is None:
            return ""

        return re.sub(
            r"[^a-z0-9]+",
            "",
            str(value).strip().lower()
        )

    def validate_ece_workbook_template(self):

        if not self.workbook.sheetnames:
            raise ValueError("ECE master workbook has no sheets.")

        sheet = self.workbook.worksheets[0]

        expected_headers = [
            "S. No.",
            "Reg. No.",
            "Name",
            "Deapartment",
            "Personal email id",
            "Phone number",
        ]

        actual_headers = [
            sheet.cell(2, column).value
            for column in range(1, 7)
        ]

        expected_normalized = [
            self.normalize_header(header)
            for header in expected_headers
        ]

        actual_normalized = [
            self.normalize_header(header)
            for header in actual_headers
        ]

        if actual_normalized != expected_normalized:
            raise ValueError(
                "ECE master workbook must have the student detail headers "
                "in row 2: S. No., Reg. No., Name, Deapartment, Personal email id, Phone number."
            )

    # ==========================================================
    # Date Helpers
    # ==========================================================

    def parse_assessment_date(self, assessment_date):

        if isinstance(assessment_date, datetime):
            return assessment_date

        if hasattr(assessment_date, "strftime"):
            return assessment_date

        text = str(assessment_date).strip()

        try:
            return datetime.strptime(text, "%d-%m-%Y")
        except ValueError:
            # Some uploaded reports use a numeric assessment label instead of a date.
            # Preserve the original value so the workbook can still be generated.
            return text

    def format_assessment_date(self, value):

        if value is None:
            return None

        if isinstance(value, datetime):
            return value.strftime("%d-%m-%Y")

        if hasattr(value, "strftime"):
            return value.strftime("%d-%m-%Y")

        text = str(value).strip()

        try:
            return datetime.strptime(text, "%d-%m-%Y").strftime("%d-%m-%Y")
        except ValueError:
            return text

    # ==========================================================
    # Backup
    # ==========================================================

    def create_backup(self):

        BACKUP_DIR.mkdir(
            parents=True,
            exist_ok=True
        )

        backup_name = datetime.now().strftime(
            "Master_Backup_%Y%m%d_%H%M%S.xlsx"
        )

        backup_path = BACKUP_DIR / backup_name

        shutil.copy2(
            self.workbook_path,
            backup_path
        )

        return backup_path

    # ==========================================================
    # Register Mapping
    # ==========================================================

    def build_register_map(self, sheet, start_row=DATA_START_ROW):

        register_map = {}

        for row in range(
            start_row,
            sheet.max_row + 1
        ):

            reg = sheet.cell(
                row=row,
                column=REGISTER_COLUMN_INDEX
            ).value

            if reg is None:
                continue

            reg = str(reg).strip()

            if reg == "":
                continue

            register_map[reg] = row

        return register_map

    # ==========================================================
    # Build All Maps
    # ==========================================================

    def build_register_maps(self):

        if self.is_ece:
            self.main_map = self.build_register_map(
                self.main_sheet,
                start_row=3
            )

            return

        self.advanced_map = self.build_register_map(
            self.advanced_sheet
        )

        self.intermediate_map = self.build_register_map(
            self.intermediate_sheet
        )

    # ==========================================================
    # Copy Cell Style
    # ==========================================================

    def copy_cell_style(
        self,
        source_cell,
        target_cell
    ):

        if not source_cell.has_style:
            return

        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    # ==========================================================
    # Copy Column Formatting
    # ==========================================================

    def copy_column_format(
        self,
        sheet,
        source_column,
        target_column
    ):

        for row in range(
            1,
            sheet.max_row + 1
        ):

            self.copy_cell_style(

                sheet.cell(row, source_column),

                sheet.cell(row, target_column)

            )

        src_letter = openpyxl.utils.get_column_letter(
            source_column
        )

        dst_letter = openpyxl.utils.get_column_letter(
            target_column
        )

        sheet.column_dimensions[
            dst_letter
        ].width = sheet.column_dimensions[
            src_letter
        ].width

    # ==========================================================
    # Find Last Assessment Block
    # ==========================================================

    def find_last_assessment_column(
        self,
        sheet
    ):

        last = FIRST_ASSESSMENT_COLUMN - DATE_BLOCK_SIZE

        for merged in sheet.merged_cells.ranges:

            if merged.min_row != DATE_ROW:
                continue

            if merged.max_col > last:
                last = merged.max_col

        return last

    # ==========================================================
    # Find Existing Date
    # ==========================================================

    def find_existing_date(
        self,
        sheet,
        assessment_date
    ):

        target = self.format_assessment_date(assessment_date)

        for merged in sheet.merged_cells.ranges:

            if merged.min_row != DATE_ROW:
                continue

            value = sheet.cell(
                merged.min_row,
                merged.min_col
            ).value

            if value is None:
                continue

            if self.format_assessment_date(value) == target:
                return merged.min_col

        return None

    # ==========================================================
    # Find Blank Assessment Block
    # ==========================================================

    def find_blank_assessment_column(self, sheet):

        blank_columns = []

        for merged in sheet.merged_cells.ranges:

            if merged.min_row != DATE_ROW:
                continue

            if merged.min_col < FIRST_ASSESSMENT_COLUMN:
                continue

            value = sheet.cell(
                merged.min_row,
                merged.min_col
            ).value

            if value is None or str(value).strip() == "":
                blank_columns.append(merged.min_col)

        if blank_columns:
            return min(blank_columns)

        return None
    
    # ==========================================================
    # Get Next Assessment Start Column
    # ==========================================================

    def get_next_assessment_column(self, sheet):

        last_column = self.find_last_assessment_column(sheet)

        if last_column < FIRST_ASSESSMENT_COLUMN:
            return FIRST_ASSESSMENT_COLUMN

        return last_column + 1


    # ==========================================================
    # Clear Assessment Block Values
    # ==========================================================

    def clear_assessment_block(
        self,
        sheet,
        start_column
    ):

        for row in range(DATA_START_ROW, sheet.max_row + 1):

            for offset in range(DATE_BLOCK_SIZE):

                sheet.cell(
                    row=row,
                    column=start_column + offset
                ).value = None


    # ==========================================================
    # Copy Assessment Block
    # ==========================================================

    def copy_assessment_block(
        self,
        sheet,
        source_start_column,
        target_start_column
    ):

        # Copy all 5 columns
        for offset in range(DATE_BLOCK_SIZE):

            source_col = source_start_column + offset
            target_col = target_start_column + offset

            self.copy_column_format(
                sheet,
                source_col,
                target_col
            )

        # Remove merged cells if already present
        remove_ranges = []

        for merged in sheet.merged_cells.ranges:

            if (
                merged.min_row == DATE_ROW
                and merged.min_col == target_start_column
            ):
                remove_ranges.append(str(merged))

        for rng in remove_ranges:
            sheet.unmerge_cells(rng)

        # Merge new date header
        sheet.merge_cells(
            start_row=DATE_ROW,
            start_column=target_start_column,
            end_row=DATE_ROW,
            end_column=target_start_column + DATE_BLOCK_SIZE - 1
        )

        # Clear copied student values
        self.clear_assessment_block(
            sheet,
            target_start_column
        )


    # ==========================================================
    # Create / Get Assessment Block
    # ==========================================================

    def create_assessment_block(
        self,
        sheet,
        assessment_date
    ):

        if self.is_ece:
            return self.create_ece_assessment_block(
                sheet,
                assessment_date
            )

        existing = self.find_existing_date(
            sheet,
            assessment_date
        )

        if existing:
            return existing

        blank_column = self.find_blank_assessment_column(
            sheet
        )

        if blank_column:

            sheet.cell(
                row=DATE_ROW,
                column=blank_column
            ).value = self.parse_assessment_date(assessment_date)

            headers = [

                SCORE_COLUMN,
                TOTAL_COLUMN,
                PERCENTAGE_COLUMN,
                RANK_COLUMN,
                STATUS_COLUMN

            ]

            for index, header in enumerate(headers):

                sheet.cell(
                    row=HEADER_ROW,
                    column=blank_column + index
                ).value = header

            return blank_column

        start_column = self.get_next_assessment_column(
            sheet
        )

        # Previous assessment block
        source_column = start_column - DATE_BLOCK_SIZE

        self.copy_assessment_block(
            sheet,
            source_column,
            start_column
        )

        # Date Header
        sheet.cell(
            row=DATE_ROW,
            column=start_column
        ).value = self.parse_assessment_date(assessment_date)

        headers = [

            SCORE_COLUMN,

            TOTAL_COLUMN,

            PERCENTAGE_COLUMN,

            RANK_COLUMN,

            STATUS_COLUMN

        ]

        for index, header in enumerate(headers):

            sheet.cell(
                row=HEADER_ROW,
                column=start_column + index
            ).value = header

        return start_column

    def create_ece_assessment_block(
        self,
        sheet,
        assessment_date
    ):

        existing = self.find_existing_date(
            sheet,
            assessment_date
        )

        if existing:
            return existing

        headers = [
            SCORE_COLUMN,
            TOTAL_COLUMN,
            PERCENTAGE_COLUMN,
            RANK_COLUMN,
            STATUS_COLUMN
        ]

        last_column = self.find_last_assessment_column(sheet)

        if last_column < FIRST_ASSESSMENT_COLUMN:

            start_column = FIRST_ASSESSMENT_COLUMN

            sheet.merge_cells(
                start_row=1,
                start_column=start_column,
                end_row=1,
                end_column=start_column + DATE_BLOCK_SIZE - 1
            )

            sheet.cell(
                row=1,
                column=start_column
            ).value = self.parse_assessment_date(assessment_date)

            for index, header in enumerate(headers):

                sheet.cell(
                    row=2,
                    column=start_column + index
                ).value = header

            return start_column

        start_column = last_column + 1

        source_column = start_column - DATE_BLOCK_SIZE

        self.copy_assessment_block(
            sheet,
            source_column,
            start_column
        )

        sheet.cell(
            row=1,
            column=start_column
        ).value = self.parse_assessment_date(assessment_date)

        for index, header in enumerate(headers):

            sheet.cell(
                row=2,
                column=start_column + index
            ).value = header

        return start_column


    # ==========================================================
    # Prepare Workbook
    # ==========================================================

    def prepare_workbook(
        self,
        assessment_date
    ):

        if self.is_ece:
            return self.create_assessment_block(
                self.main_sheet,
                assessment_date
            )

        advanced_column = self.create_assessment_block(
            self.advanced_sheet,
            assessment_date
        )

        intermediate_column = self.create_assessment_block(
            self.intermediate_sheet,
            assessment_date
        )

        return advanced_column, intermediate_column


    # ==========================================================
    # Build Uploaded Data Dictionary
    # ==========================================================

    def build_uploaded_dictionary(self, dataframe):

        uploaded = {}

        for _, row in dataframe.iterrows():

            reg = clean_register_number(
                row[REGISTER_COLUMN]
            ).strip()

            if not reg:
                continue

            uploaded[reg] = {

                "score": row[SCORE_COLUMN],

                "total": row[TOTAL_COLUMN],

                "department": row.get(DEPARTMENT_COLUMN, ""),

                "percentage": row[PERCENTAGE_COLUMN],

                "rank": row[RANK_COLUMN],

                "status": row[STATUS_COLUMN]

            }

        return uploaded


    # ==========================================================
    # Update One Sheet
    # ==========================================================

    def update_sheet(
            self,
            sheet,
            register_map,
            dataframe,
            start_column
    ):

        uploaded = self.build_uploaded_dictionary(
            dataframe
        )

        unknown = sorted(
            reg
            for reg in uploaded
            if reg not in register_map
        )

        updated = 0
        absent = 0

        for reg, excel_row in register_map.items():

            student = uploaded.get(reg)

            # -----------------------------
            # Student Present
            # -----------------------------

            if student:

                sheet.cell(
                    excel_row,
                    start_column
                ).value = student["score"]

                sheet.cell(
                    excel_row,
                    start_column + 1
                ).value = student["total"]

                sheet.cell(
                    excel_row,
                    start_column + 2
                ).value = student["percentage"]

                sheet.cell(
                    excel_row,
                    start_column + 3
                ).value = student["rank"]

                sheet.cell(
                    excel_row,
                    start_column + 4
                ).value = student["status"]

                updated += 1

            # -----------------------------
            # Student Absent
            # -----------------------------

            else:

                sheet.cell(
                    excel_row,
                    start_column
                ).value = 0

                sheet.cell(
                    excel_row,
                    start_column + 1
                ).value = 0

                sheet.cell(
                    excel_row,
                    start_column + 2
                ).value = 0

                sheet.cell(
                    excel_row,
                    start_column + 3
                ).value = 0

                sheet.cell(
                    excel_row,
                    start_column + 4
                ).value = ABSENT

                absent += 1

        return {

            "updated": updated,

            "absent": absent,

            "total": len(register_map),

            "skipped": len(unknown),

            "missing_registers": unknown

        }


    # ==========================================================
    # Update Advanced Sheet
    # ==========================================================

    def update_advanced_sheet(
            self,
            dataframe,
            start_column
    ):

        return self.update_sheet(

            self.advanced_sheet,

            self.advanced_map,

            dataframe,

            start_column

        )


    # ==========================================================
    # Update Intermediate Sheet
    # ==========================================================

    def update_intermediate_sheet(
            self,
            dataframe,
            start_column
    ):

        return self.update_sheet(

            self.intermediate_sheet,

            self.intermediate_map,

            dataframe,

            start_column

        )

    def update_ece_sheet(
            self,
            dataframe,
            start_column
    ):

        return self.update_sheet(

            self.main_sheet,

            self.main_map,

            dataframe,

            start_column

        )


    # ==========================================================
    # Update Workbook
    # ==========================================================

    def update_workbook(

            self,

            advanced_df,

            intermediate_df,

            assessment_date

    ):

        if self.is_ece:

            start_column = self.prepare_workbook(
                assessment_date
            )

            self.build_register_maps()

            ece_summary = self.update_ece_sheet(

                advanced_df,

                start_column

            )

            return {
                "ECE": ece_summary
            }

        advanced_column, intermediate_column = \
            self.prepare_workbook(
                assessment_date
            )

        self.build_register_maps()

        advanced_summary = self.update_advanced_sheet(

            advanced_df,

            advanced_column

        )

        intermediate_summary = self.update_intermediate_sheet(

            intermediate_df,

            intermediate_column

        )

        return {

            "Advanced": advanced_summary,

            "Intermediate": intermediate_summary

        }
    
    # ==========================================================
    # Clear Performer Block
    # ==========================================================

    def clear_performer_block(self, sheet, start_row):

        for row in range(start_row, start_row + 10):

            for column in range(2, 16):

                sheet.cell(
                    row=row,
                    column=column
                ).value = None


    # ==========================================================
    # Write Performer Block
    # ==========================================================

    def write_performer_block(
            self,
            sheet,
            start_row,
            label,
            dataframe,
            top=True
    ):

        if sheet is None:
            return

        block = dataframe.copy()

        if STATUS_COLUMN in block.columns:
            block = block[
                block[STATUS_COLUMN] == ATTENDED
            ]

        if RANK_COLUMN in block.columns:

            if top:
                block = block.nsmallest(10, RANK_COLUMN)
            else:
                block = block.nlargest(10, RANK_COLUMN)

        elif SCORE_COLUMN in block.columns:

            sort_columns = [SCORE_COLUMN]
            ascending = [False]

            if TIME_COLUMN in block.columns:
                sort_columns.append(TIME_COLUMN)
                ascending.append(True)

            block = block.sort_values(
                by=sort_columns,
                ascending=ascending
            ).head(10)

        else:
            block = block.head(10)

        self.clear_performer_block(
            sheet,
            start_row
        )

        sheet.cell(
            row=start_row,
            column=1
        ).value = label

        for offset, (_, student) in enumerate(block.iterrows()):

            row = start_row + offset

            sheet.cell(row, 2).value = student[REGISTER_COLUMN]
            sheet.cell(row, 3).value = student[NAME_COLUMN]
            sheet.cell(row, 4).value = student.get(DEPARTMENT_COLUMN, "")
            sheet.cell(row, 5).value = student[SCORE_COLUMN]
            sheet.cell(row, 6).value = student[TOTAL_COLUMN]
            sheet.cell(row, 7).value = student[PERCENTAGE_COLUMN]
            sheet.cell(row, 8).value = student[RANK_COLUMN]

    # ==========================================================
    # Update Top 10 Performers
    # ==========================================================

    def update_top10(
            self,
            assessment_date,
            advanced_df,
            intermediate_df
    ):

        if self.is_ece:
            return self.update_top10_ece(
                assessment_date,
                advanced_df
            )

        if self.top_sheet is None:
            return

        date_value = self.parse_assessment_date(
            assessment_date
        )

        self.top_sheet.cell(1, 2).value = date_value
        self.top_sheet.cell(1, 9).value = date_value

        self.write_performer_block(
            self.top_sheet,
            3,
            "Advanced",
            advanced_df,
            top=True
        )

        self.write_performer_block(
            self.top_sheet,
            14,
            "Intermediate",
            intermediate_df,
            top=True
        )

    def update_top10_ece(self, assessment_date, dataframe):

        if self.top_sheet is None:
            return

        date_value = self.parse_assessment_date(
            assessment_date
        )

        self.top_sheet.cell(1, 2).value = date_value

        self.write_performer_block(
            self.top_sheet,
            3,
            "ECE",
            dataframe,
            top=True
        )

    # ==========================================================
    # Update Last 10 Performers
    # ==========================================================

    def update_last10(
            self,
            assessment_date,
            advanced_df,
            intermediate_df
    ):

        if self.is_ece:
            return self.update_last10_ece(
                assessment_date,
                advanced_df
            )

        if self.last_sheet is None:
            return

        date_value = self.parse_assessment_date(
            assessment_date
        )

        self.last_sheet.cell(1, 2).value = date_value
        self.last_sheet.cell(1, 9).value = date_value

        self.write_performer_block(
            self.last_sheet,
            3,
            "Advanced",
            advanced_df,
            top=False
        )

        self.write_performer_block(
            self.last_sheet,
            14,
            "Intermediate",
            intermediate_df,
            top=False
        )

    def update_last10_ece(self, assessment_date, dataframe):

        if self.last_sheet is None:
            return

        date_value = self.parse_assessment_date(
            assessment_date
        )

        self.last_sheet.cell(1, 2).value = date_value

        self.write_performer_block(
            self.last_sheet,
            3,
            "ECE",
            dataframe,
            top=False
        )


    # ==========================================================
    # Save Workbook
    # ==========================================================

    def save_workbook(
            self,
            output_path
    ):

        output_path = Path(output_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        self.workbook.save(
            output_path
        )

        return output_path


    # ==========================================================
    # Complete Processing
    # ==========================================================

    def process(

            self,

            advanced_dataframe,

            intermediate_dataframe,

            assessment_date,

            output_file

    ):

        self.load_workbook()

        self.create_backup()

        if self.is_ece:

            summary = self.update_workbook(

                advanced_dataframe,

                intermediate_dataframe,

                assessment_date

            )

            self.update_top10(

                assessment_date,

                advanced_dataframe,

                intermediate_dataframe

            )

            self.update_last10(

                assessment_date,

                advanced_dataframe,

                intermediate_dataframe

            )

            self.save_workbook(

                output_file

            )

            return summary

        summary = self.update_workbook(

            advanced_dataframe,

            intermediate_dataframe,

            assessment_date

        )

        self.update_top10(

            assessment_date,

            advanced_dataframe,

            intermediate_dataframe

        )

        self.update_last10(

            assessment_date,

            advanced_dataframe,

            intermediate_dataframe

        )

        self.save_workbook(

            output_file

        )

        return summary
